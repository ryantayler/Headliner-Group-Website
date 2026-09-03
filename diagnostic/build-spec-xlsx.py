#!/usr/bin/env python3
"""Build the editable spec workbook from content.js.

content.js is the single source. This script never invents content, it only
lays it out for editing. Run it again after any change to content.js.

    python3 build-spec-xlsx.py

Colour key used throughout, and explained on the Read me tab.
  Yellow  edit freely, this is wording that reaches the reader
  Grey    do not edit, the engine matches on these ids
  White   numbers and settings, safe to change, see the Thresholds tab
"""
import json, re, sys, collections
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HERE = Path(__file__).parent
OUT = HERE / "Business-Diagnostic-Spec.xlsx"

src = (HERE / "content.js").read_text()
i = src.index("window.DIAG =")
D = json.loads(src[i + len("window.DIAG ="):].strip()[:-1], object_pairs_hook=collections.OrderedDict)

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="1A1A1A")
EDIT_FILL = PatternFill("solid", fgColor="FFF6D8")   # yellow, wording
LOCK_FILL = PatternFill("solid", fgColor="EDEDED")   # grey, ids
PINK = "C4174F"
THIN = Side(style="thin", color="D5D5D5")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

SLOT_RE = re.compile(r"\{(q\d+)\.(band|exact)\}|\{d\.(\w+)\}")
DERIVED_NOTE = {"primaryShort": "the primary constraint's short name",
                "primaryName": "the primary constraint's full name",
                "unownedLayers": "the sentence listing which layers sit with the owner"}


def slots_in(*texts):
    texts = [x for t in texts for x in (t if isinstance(t, list) else [t])]
    """Every variable used in these cells, written so somebody editing can see what
    they have to work with. Blank when the copy carries no data."""
    found = []
    for t in texts:
        for m in SLOT_RE.finditer(t or ""):
            if m.group(3):
                tag = "{d." + m.group(3) + "}"
            else:
                tag = "{" + m.group(1) + "." + m.group(2) + "}"
            if tag not in found:
                found.append(tag)
    return ", ".join(found)



wb = Workbook()
wb.remove(wb.active)


def sheet(name, headers, rows, widths, edit_cols=(), lock_cols=(), wrap_cols=(), note=None):
    ws = wb.create_sheet(name[:31])
    r = 1
    if note:
        ws.cell(1, 1, note).font = Font(name=FONT, size=10, italic=True, color="5B6360")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
        ws.cell(1, 1).alignment = Alignment(vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 30
        r = 2
    for c, h in enumerate(headers, 1):
        cell = ws.cell(r, c, h)
        cell.font = Font(name=FONT, size=10, bold=True, color="FFFFFF")
        cell.fill = HEAD_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = BORDER
    ws.row_dimensions[r].height = 30
    for row in rows:
        r += 1
        for c, v in enumerate(row, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name=FONT, size=10)
            cell.border = BORDER
            cell.alignment = Alignment(vertical="top", wrap_text=(c in wrap_cols))
            if c in edit_cols:
                cell.fill = EDIT_FILL
            elif c in lock_cols:
                cell.fill = LOCK_FILL
                cell.font = Font(name=FONT, size=9, color="6B7078")
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = ws.cell(r if False else (2 if not note else 3), 1)
    ws.auto_filter.ref = f"A{2 if note else 1}:{get_column_letter(len(headers))}{r}"
    return ws


# ---------------------------------------------------------------- 1. Read me
readme = [
    ["What this is", "The full content and logic of the diagnostic tool, laid out so you can rewrite the wording. Every question, every answer option, every line of the report."],
    ["How to use it", "Edit the yellow cells. Leave the grey ones alone. Hand the file back and the tool gets rebuilt from it."],
    ["Yellow cells", "Wording that reaches the reader. Rewrite freely."],
    ["Grey cells", "Ids the engine matches on, and columns it works out for you. Changing one breaks the link between a question, its answer and the report line that quotes it."],
    ["Variables", "Anything in {curly braces} is filled in from the answers. Write band phrases lower case, because sentence case is applied after a variable is filled in, so one landing at the start of a sentence gets capitalised for you."],
    ["White cells", "Numbers and settings. Safe to change, and the Thresholds tab explains what each one does."],
    ["", ""],
    ["The one rule", "The report names one constraint. Not a list. If an edit starts adding a second finding, it is working against the tool."],
    ["The second rule", "No scores, no grades, no dollar figures, no timeframes. None of those can be defended from banded answers."],
    ["Voice", "No dashes of any kind. No colons except before a list. Australian spelling. Contractions throughout."],
    ["", ""],
    ["Tabs, in order", ""],
    ["Questions", "All 52 questions. Wording, help text, which block each one feeds, how heavily it counts."],
    ["Answer options", "Every option under every question. This is where the bands live, and the band phrase is what gets quoted in the report."],
    ["Constraints and order", "The seven constraints, the fixed order they are checked in, and what suppresses what."],
    ["Report, constraint", "The constraint blocks. Opening line, evidence clauses, closing line, and the fix actions."],
    ["Report, minor", "The one line that prints when a second constraint is loud but downstream."],
    ["Report, risk flags", "All 16 risk flags. Definition, the version used when the answer was Not sure, and the fix."],
    ["Report, risk section", "The line that opens the risk section."],
    ["Report, don't do yet", "The closing section. What not to touch until the constraint is fixed."],
    ["Report, open and close", "The opening and closing of the report, including the three opening variants."],
    ["Thresholds", "Every tunable number, and what moving it does."],
    ["Logic", "How the engine picks the primary, the minor, and the risks. Read this before changing a threshold."],
    ["Data slots", "Every variable, every phrase it can print, and which report blocks quote it. Check here before changing a band phrase, because it shows you the sentences that band lands in."],
]
ws = sheet("Read me", ["", ""], readme, [26, 118], edit_cols=(), wrap_cols=(2,))
for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=1):
    row[0].font = Font(name=FONT, size=10, bold=True, color=PINK)
ws.auto_filter.ref = None
ws.freeze_panes = None

# ------------------------------------------------------------- 2. Questions
SECTION_LABEL = {"profiling": "Profiling, calibrates the bands", "risk": "Risk profile"}
rows = []
for q in D["questions"]:
    sec = D["constraints"].get(q["section"], {}).get("name") or SECTION_LABEL.get(q["section"], q["section"])
    rows.append([
        q["n"], q["id"], sec, q["type"],
        q["text"], q.get("help", ""),
        q.get("weight", 0),
        (q["exact"]["label"] + "  (" + q["exact"]["unit"] + ")") if q.get("exact") else "",
        q.get("scoreRule", ""),
        ("Only shown when " + q["showIf"]["q"] + " is answered with something other than "
         + "/".join(q["showIf"].get("notOnly", []))) if q.get("showIf") else "",
    ])
sheet("Questions",
      ["#", "id", "Feeds", "Format", "Question wording", "Help text under the question",
       "Weight in its block", "Optional exact figure prompt", "Special scoring", "Only shown if"],
      rows, [5, 7, 26, 9, 58, 44, 10, 40, 16, 40],
      edit_cols=(5, 6, 8), lock_cols=(2, 10), wrap_cols=(3, 5, 6, 8, 10),
      note="Weight is how heavily a question counts inside its own constraint block. 0 means it does not score at all, it only raises risk flags or sets the profile. Profiling and risk questions never score a constraint.")

# --------------------------------------------------------- 3. Answer options
rows = []
for q in D["questions"]:
    for o in q["options"]:
        flags = []
        for f in o.get("flags", []):
            flags.append(f if isinstance(f, str) else f"{f['id']} ({f['sev']})")
        notes = []
        if o.get("notSure"): notes.append("Not sure, scores nothing and feeds data blindness")
        if o.get("exclusive"): notes.append("Clears every other tick")
        if o.get("disqualify"): notes.append(f"Rules out {o['disqualify']} entirely")
        if o.get("oneoff"): notes.append("Shifts the value block onto referral")
        if o.get("boost"): notes.append(f"Promotes {o['boost']} to the top severity")
        rows.append([
            q["n"], q["id"], q["text"][:58] + ("..." if len(q["text"]) > 58 else ""),
            o["id"], o["text"], o.get("band", ""),
            o.get("w", "") if "w" in o else "", ", ".join(flags), "; ".join(notes),
        ])
sheet("Answer options",
      ["Q#", "q id", "Question", "opt", "Answer wording", "Band phrase, as quoted in the report",
       "Severity 0 to 100", "Risk flags raised (severity)", "Behaviour"],
      rows, [5, 7, 44, 5, 46, 46, 11, 34, 34],
      edit_cols=(5, 6), lock_cols=(2, 4), wrap_cols=(3, 5, 6, 8, 9),
      note="The band phrase is dropped into report sentences, so it has to read inside one. \"in the 10 to 25% range\" works, \"10-25%\" does not. Leave it empty on a Not sure option, which is what makes that clause drop out of the sentence instead of printing a hole.")

# ------------------------------------------------- 4. Constraints and order
rows = []
DEFN = {
 "cashflow":"Profitable but illiquid. The work is sold, the money hasn't landed, and there's nothing to fund wages, stock or gear. A timing problem, not a rate problem.",
 "talent":"A missing layer, not missing hands. The role that should exist doesn't, and the owner is absorbing it. Hiring another doer doesn't fix it.",
 "fulfilment":"Capacity, not capability. They know how to do the work and need more hours of it. Blown lead times, waitlists, turning jobs away, quality slipping.",
 "value":"They buy, they don't perceive enough value to stay. Filling a leaky bucket. The customer's perception, nothing to do with pricing or margin.",
 "offer":"Enquiries arrive, the offer doesn't land. Covers offering the wrong thing entirely, not just weak sales technique.",
 "demand":"Spare capacity sitting idle. Could take on more tomorrow, nobody's asking.",
 "margin":"Busy, full, growing, nothing left at the end. Profit per unit is short."}
WHY = {
 "cashflow":"First, because nothing else can be actioned without money to fund it.",
 "talent":"Above fulfilment, because capacity added under a missing layer just loads the layer.",
 "fulfilment":"Above demand, because adding demand to a business that can't deliver accelerates the damage.",
 "value":"Above offer, because winning more customers into a leaky bucket wastes the win.",
 "offer":"Above demand, because more enquiries against an offer that doesn't land changes nothing.",
 "demand":"Second last. Real, but only once everything above it can absorb the work.",
 "margin":"Last, because it is usually a symptom of something above it rather than a standalone constraint."}
for n, cid in enumerate(D["chain"], 1):
    supp = D["suppresses"].get(cid, [])
    ht = D["hardTriggers"].get(cid, [])
    trig = []
    for rule in ht:
        trig.append(" AND ".join(f"{p[0]} is {'/'.join(p[1])}" for p in rule["all"]))
    rows.append([n, cid, D["constraints"][cid]["name"], DEFN[cid], WHY[cid],
                 ", ".join(supp) or "nothing", " OR ".join(trig)])
sheet("Constraints and order",
      ["Order", "id", "Name", "Definition", "Why it sits here", "Directly suppresses", "Hard trigger, fires regardless of score"],
      rows, [6, 11, 24, 62, 56, 18, 44],
      edit_cols=(3, 4), lock_cols=(2,), wrap_cols=(4, 5, 6, 7),
      note="The engine walks this list top to bottom and calls the first one that fails. It never picks the highest scoring one. Suppression runs through the chain, so talent suppresses fulfilment and everything fulfilment suppresses underneath it.")

# ------------------------------------------------ 5. Report, constraint blocks
rows = []
for cid in D["chain"]:
    b = D["blocks"]["constraintDef"][cid]
    f = D["blocks"]["constraintFix"][cid]
    rows.append([cid, "Title", b["title"], "", ""])
    rows.append([cid, "Title, nothing failing", b["titleLoose"], "", ""])
    opens = b["open"] if isinstance(b["open"], list) else [b["open"]]
    for n, o in enumerate(opens, 1):
        rows.append([cid, "Opening" if len(opens) == 1 else f"Opening sentence {n}", o, "", slots_in(o)])
    for n, e in enumerate(b.get("evidence", []), 1):
        rows.append([cid, f"Evidence clause {n}", e["banded"], e.get("precise", ""), slots_in(e["banded"], e.get("precise"))])
    rows.append([cid, "Closing", b["close"], "", slots_in(b["close"])])
    rows.append([cid, "Fix, lead line", f["lead"], "", ""])
    for n, a in enumerate(f["actions"], 1):
        cond = "  ".join(f"{p[0]} is {'/'.join(p[1])}" for p in a.get("when", []))
        rows.append([cid, f"Fix action {n}", a["text"], "", cond or "always prints"])
sheet("Report, constraint",
      ["Constraint", "Part", "Banded version, always used", "Precise version, used only when they typed the figure", "Variables, or when a fix action prints"],
      rows, [13, 22, 82, 68, 26],
      edit_cols=(3, 4), lock_cols=(1, 2, 5), wrap_cols=(3, 4, 5),
      note="Evidence clauses are joined into one sentence with commas and a final \"and\". Write each one lower case and without a full stop. A clause whose data came back Not sure is dropped from the sentence rather than printed empty, which is why they are separate rows.")

# ------------------------------------------------------- 6. Report, minor
rows = [[cid, D["constraints"][cid]["name"], D["blocks"]["minorConstraint"][cid], slots_in(D["blocks"]["minorConstraint"][cid])] for cid in D["chain"]]
sheet("Report, minor", ["id", "Constraint", "The one line printed when this is a loud but downstream second finding", "Variables in this row"],
      rows, [13, 24, 100, 24], edit_cols=(3,), lock_cols=(1, 4), wrap_cols=(3, 4),
      note="One line only, no fix and no detail. Naming it proves the tool knows the difference between a cause and a shadow, and a full fix section would undo that. {d.primaryShort} resolves to the primary constraint's short name.")

# --------------------------------------------------- 7. Report, risk flags
rows = []
for fid, meta in D["flags"].items():
    rd = D["blocks"]["riskDef"][fid]
    fixes = D["blocks"]["riskFix"].get(fid, [])
    rows.append([meta["family"], fid, meta["name"], rd.get("banded", ""), rd.get("precise", ""), rd.get("alt", ""),
                 fixes[0] if len(fixes) > 0 else "", fixes[1] if len(fixes) > 1 else "",
                 slots_in(rd.get("banded"), rd.get("precise"))])
sheet("Report, risk flags",
      ["Family", "id", "Name", "Definition", "Precise version", "Version used when the answer was Not sure", "What to do", "Variables in this row"],
      rows, [15, 19, 22, 72, 56, 72, 66, 24],
      edit_cols=(3, 4, 5, 6, 7), lock_cols=(2, 8), wrap_cols=(4, 5, 6, 7, 8),
      note="At most three flags print, from the top scoring family only. A flag has to clear the print threshold on the Thresholds tab to be named at all.")

# ------------------------------------------------- 8. Report, risk section
rows = [["Section lead, printed once above the risks", D["blocks"]["riskLead"]]]
sheet("Report, risk section", ["Part", "Wording"], rows, [42, 108],
      edit_cols=(2,), wrap_cols=(2,),
      note="Risks print as one list, loudest first, regardless of which family they came from. Families still weight the internal ordering, they no longer split the section. The old family framing block, the separate minor risk section and the compound blocks were all removed after they read as filler in testing.")

# ------------------------------------------------ 10. Report, don't do yet
rows = []
for cid in D["chain"]:
    b = D["blocks"]["dontDoYet"][cid]
    rows.append([cid, "Lead line", b["lead"]])
    for n, it in enumerate(b["items"], 1):
        rows.append([cid, f"Item {n}", it])
sheet("Report, don't do yet", ["Constraint", "Part", "Wording"],
      rows, [13, 12, 122], edit_cols=(3,), lock_cols=(1, 2), wrap_cols=(3,),
      note="Telling an owner to stop doing the wrong thing is usually worth more than another action item, because they are already mid action on it. This section does not print when nothing is failing.")

# --------------------------------------------- 11. Report, open and close
rows = [
    ["Opening", "normal", "Used whenever a constraint actually failed", D["blocks"]["opening"]["normal"]],
    ["Opening", "wellRun", "Nothing cleared the floor. Nothing is failing", D["blocks"]["opening"]["wellRun"]],
    ["Opening", "noneSevere", "Nothing failed outright but something is tight", D["blocks"]["opening"]["noneSevere"]],
    ["Opening", "tooUnsure", "Too many Not sure answers to stand behind a diagnosis", D["blocks"]["opening"]["tooUnsure"]],
    ["Opening", "privacy", "Small print under the opening, on every report", D["blocks"]["opening"]["privacy"]],
    ["Heading", "titleUnsure", "Replaces the constraint heading when there is not enough to call it", D["blocks"]["titleUnsure"]],
    ["Body", "looseBody", "Replaces the constraint block when nothing is failing", D["blocks"]["looseBody"]],
    ["Body", "unsureBody", "Replaces the constraint block when there were too many Not sure answers", D["blocks"]["unsureBody"]],
    ["Closing", "text", "Normal close", D["blocks"]["closing"]["text"]],
    ["Closing", "loose", "Close used when nothing is failing", D["blocks"]["closing"].get("loose", "")],
    ["Closing", "unsure", "Close used when there were too many Not sure answers", D["blocks"]["closing"].get("unsure", "")],
    ["Closing", "cta", "Call to action. Deliberately empty, still an open question", D["blocks"]["closing"]["cta"]],
]
rows = [r + [slots_in(r[3])] for r in rows]
sheet("Report, open and close", ["Section", "Variant", "When it is used", "Wording", "Variables in this row"],
      rows, [11, 13, 48, 98, 24], edit_cols=(4,), lock_cols=(2, 5), wrap_cols=(3, 4, 5),
      note="Blank line between paragraphs is written as two newlines in the source. Keep paragraphs short.")

# ------------------------------------------------------- 12. Thresholds
TH = {
 "PRIMARY_FAIL": "A constraint fails, and gets called, at or above this score. Raise it and fewer businesses get a hard finding. Lower it and the first thing in the chain wins too easily.",
 "MINOR_PRINT": "A second constraint has to reach this before it prints as a minor. Deliberately higher than PRIMARY_FAIL, so a minor has to be louder than the bar that would have made it a primary in its own right.",
 "FALLBACK_FLOOR": "When nothing failed, anything under this counts as a well run business and the report switches to the softer wording. Above it, the report says nothing is failing but here is the tightest thing.",
 "FLAG_PRINT": "An individual flag has to reach this before it is named. Below it the flag still counts toward its family score.",
 "MAX_FLAGS_SHOWN": "Hard cap on how many risks get named, across all three families. Four keeps the report short and certain.",
 "MAX_ACTIONS": "Hard cap on how many fix actions print for the constraint. Conditional actions that did not fire never count toward it.",
 "NOT_SURE_DATA_FLAG": "This many Not sure answers raises data blindness on its own, regardless of which questions they were.",
}
rows = [[k, D["thresholds"][k], TH[k]] for k in D["thresholds"]]
sheet("Thresholds", ["Setting", "Value", "What moving it does"], rows, [24, 9, 118],
      lock_cols=(1,), wrap_cols=(3,),
      note="Change the value column. Every one of these is a judgement call rather than a fact, and they are the first thing to tune once you have run real businesses through the tool.")

# ------------------------------------------------------------- 13. Logic
logic = [
 ["Picking the primary", "The engine walks the seven constraints in the fixed order on the Constraints tab and calls the first one that fails. A constraint fails when its hard trigger fires, or when its block score reaches PRIMARY_FAIL."],
 ["", "It never picks the highest scoring one. A downstream constraint with a higher score does not overtake a failing upstream one, because fixing downstream while upstream is broken makes things worse."],
 ["", "There is no dominance override on purpose. The protection against calling a mild problem is that the bar is set high enough that clearing it means genuinely broken."],
 ["Block scores", "Each constraint's score is the weighted average of the severity of its five answers, using the weights on the Questions tab. A Not sure answer contributes zero severity but still counts in the average, so uncertainty lowers a score rather than inflating it."],
 ["Disqualifiers", "Some answers rule a constraint out entirely. Not being profitable rules out cash flow, because cash flow constrained means profitable but illiquid. Not being able to deliver double rules out demand."],
 ["When nothing fails", "The engine names the highest scoring constraint and switches the wording. Below FALLBACK_FLOOR it prints the well run version with no evidence paragraph and no Don't do this yet section, because there is nothing to hold off on."],
 ["Picking the minor", "Every constraint other than the primary that reaches MINOR_PRINT is a candidate. Suppression then removes the ones explained by the primary, running through the chain, so talent removes fulfilment and value underneath it. The highest scoring survivor prints, capped at one."],
 ["", "Anything surviving suppression is always downstream of the primary, which is why the single line printed about it can always say so honestly."],
 ["Risk families", "Family score is 60 percent of the loudest single flag plus 40 percent of the proportion of that family's flags that cleared FLAG_PRINT. One severe flag alone is enough to print a family. Several moderate ones also lift it."],
 ["Not sure", "Every Not sure answer counts. Some raise data blindness directly. Repeated blindness compounds, and NOT_SURE_DATA_FLAG unanswerable questions raises the flag on its own. Fifteen or more switches the report opening to the provisional version."],
 ["Compound blocks", "A compound block prints when the primary constraint and a named risk flag are both present. At most one prints, the lowest priority number that matches."],
 ["The exact figure", "When somebody types a precise number it only changes the wording, never the finding. The logic already fired on the band. This keeps the engine simple and stops a typo flipping the verdict."],
 ["What the report never does", "No score, no grade, no dollar projection, no timeframe. None of those can be defended from banded inputs, and one indefensible number discredits everything around it."],
]
sheet("Logic", ["", "How it works"], logic, [24, 128], wrap_cols=(2,))

# ---------------------------------------------------------- 14. Data slots
# Where each variable is quoted, and every phrase it can print. Change a band on the
# Answer options tab and this is the list of sentences it lands in.
def block_paths(node, path=""):
    if isinstance(node, dict):
        for k, v in node.items():
            yield from block_paths(v, f"{path} / {k}" if path else k)
    elif isinstance(node, list):
        for n, v in enumerate(node, 1):
            yield from block_paths(v, f"{path} [{n}]")
    elif isinstance(node, str):
        yield path, node

USES = {}
for path, text in block_paths(D["blocks"]):
    for m in SLOT_RE.finditer(text):
        tag = "{d." + m.group(3) + "}" if m.group(3) else "{" + m.group(1) + "." + m.group(2) + "}"
        USES.setdefault(tag, []).append(path)

rows = []
for q in D["questions"]:
    band_tag, exact_tag = "{" + q["id"] + ".band}", "{" + q["id"] + ".exact}"
    if band_tag in USES:
        phrases = [o.get("band", "") for o in q["options"] if o.get("band")]
        blanks = [o["text"] for o in q["options"] if not o.get("band")]
        rows.append([band_tag, q["n"], q["text"], " | ".join(phrases),
                     "clause is dropped on: " + ", ".join(blanks) if blanks else "always resolves",
                     "  ".join(sorted(set(USES[band_tag])))])
    if q.get("exact"):
        rows.append([exact_tag, q["n"], q["text"],
                     f"the number they typed, in {q['exact']['unit']}",
                     "optional, so the precise wording only prints when it was filled in",
                     "  ".join(sorted(set(USES.get(exact_tag, [])))) or "OFFERED BUT NEVER QUOTED, so remove the prompt or use it"])
for tag, note in [("{d.primaryShort}", "the primary constraint's short name, lower case"),
                  ("{d.primaryName}", "the primary constraint's full name, lower case"),
                  ("{d.unownedLayers}", "a sentence naming which layers sit with the owner")]:
    if tag in USES:
        rows.append([tag, "", "derived, not a question", note, "always resolves", "  ".join(sorted(set(USES[tag])))])

sheet("Data slots",
      ["Variable", "Q#", "Question it comes from", "Every phrase it can print", "When it does not resolve", "Report blocks that quote it"],
      rows, [17, 5, 54, 70, 52, 60],
      lock_cols=(1,), wrap_cols=(3, 4, 5, 6),
      note="An exact figure is only offered on questions whose answer actually reaches the report prose, not on all fifty two. Asking for a number the report never uses is friction for nothing. Sentence case is applied after a variable is filled in, so write band phrases lower case and let the engine capitalise the ones that land at the start of a sentence.")

wb.save(OUT)
print(f"wrote {OUT.name}  ({len(wb.sheetnames)} tabs)")
for s in wb.sheetnames:
    print(f"  {s}")
